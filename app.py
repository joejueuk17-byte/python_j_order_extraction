import os
from bs4 import BeautifulSoup
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from datetime import datetime

def parse_html(file_path):
    if not os.path.exists(file_path):
        print(f"Error: {file_path} not found.")
        return None, None

    with open(file_path, 'r', encoding='utf-8') as f:
        soup = BeautifulSoup(f, 'html.parser')
    
    # 1. Extract Date (Removing the time portion)
    try:
        # Specifically targeting the div containing the order date text
        date_raw = soup.find('div', class_='title_f25').find_next('div', class_='align-items-center').text.strip()
        # Split by " - " and take the first part: "19 Jan 2026"
        order_date_str = date_raw.split(' - ')[0].strip()

        # Try parsing with flexible formats
        dt = parse_date_flexible(order_date_str)

        if dt:
            order_date = dt.strftime("%d %b, %Y")   # → "05 Feb, 2026"
        else:
            order_date = "Unknown Date"
    except:
        order_date = "Unknown Date"

    # 2. Extract Items
    items = []
    # Target rows inside the table that have the 'table-active' class
    product_rows = soup.find('table', class_='table-list').find('tbody').find_all('tr', class_='table-active')
    
    for row in product_rows:
        # The product name and link are inside the second <td>
        cells = row.find_all('td')
        if len(cells) < 2:
            continue
            
        details_cell = cells[1]
        link_tag = details_cell.find('a', href=True)
        name_tag = details_cell.find('h5')
        
        if link_tag and name_tag:
            name = name_tag.get_text(strip=True)
            link = "https://www.suruga-ya.com" + link_tag['href']
            
            # Extract prices
            try:
                price_cell = cells[2]
                #disc_price = price_cell.find('div', class_='price-new').text.replace('JPY', '').replace(',', '').strip()
                #orig_price = price_cell.find('div', class_='price-old').text.replace('JPY', '').replace(',', '').strip()
                disc_price = get_price(price_cell, 'price-new')
                orig_price = get_price(price_cell, 'price-old') or disc_price
                
                items.append({
                    "name": name,
                    "link": link,
                    "disc": int(disc_price),
                    "orig": int(orig_price)
                })
            except:
                continue
            
    return order_date, items
def parse_date_flexible(date_str):
    # Possible formats the site might use
    formats = [
        "%d %b %Y",
        "%d %B %Y",
        "%d %b, %Y",
        "%d %B, %Y",
        "%d %b %Y %H:%M",
        "%d %B %Y %H:%M",
        "%Y/%m/%d",
        "%Y-%m-%d",
    ]

    for fmt in formats:
        try:
            return datetime.strptime(date_str, fmt)
        except:
            pass

    return None  # No format matched

def get_price(cell, class_name):
    el = cell.find('div', class_=class_name)
    if not el:
        return None
    return el.text.replace('JPY', '').replace(',', '').strip()

def run_app():
    # --- USER INPUTS ---
    try:
        first_line = int(input("Enter [first line number] (e.g. 875): "))
        gbp_paid = float(input("Enter [total gbp paid]: ").replace(',', ''))
        yen_paid = float(input("Enter [total yen paid]: ").replace(',', ''))
    except ValueError:
        print("Invalid input. Please enter valid numbers.")
        return

    order_date, items = parse_html("order.html")
    if not items:
        print("Failed to retrieve items. Please check if 'order.html' contains the order table.")
        return

    wb = Workbook()
    ws = wb.active
    
    last_item_num = first_line + len(items) - 1
    
    # --- STYLES ---
    blue_font = Font(color="0000FF")
    strike_font = Font(strikethrough=True)
    blue_underline_font = Font(color="0000FF", underline="single")
    yellow_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
    pastel_orange_fill = PatternFill(start_color="FFCC99", end_color="FFCC99", fill_type="solid")

    for i, item in enumerate(items):
        curr = first_line + i
        
        # 1. Col A: Date
        ws.cell(row=curr, column=1, value=order_date)
        
        # 2. Col B: Hyperlinked Name
        #ws.cell(row=curr, column=2, value=f'=HYPERLINK("{item["link"]}", "{item["name"]}")')
        # good ws.cell(row=curr, column=2, value=item["name"])
        
        cell_b = ws.cell(row=curr, column=2, value=item["name"]) 
        cell_b.hyperlink = item["link"]
        cell_b.font = blue_underline_font

        # 6. Col F: Discount Price (Blue)
        cell_f = ws.cell(row=curr, column=6, value=item["disc"])
        cell_f.font = blue_font
        
        # 7. Col G: Original Price (Strikeout)
        cell_g = ws.cell(row=curr, column=7, value=item["orig"])
        cell_g.font = strike_font
        
        # 8. Col H: Formula
        ws.cell(row=curr, column=8, value=f"=$G{curr}*$Q${first_line}")

        # Summary Row (First Line)
        if curr == first_line:
            # 5. Col E
            ws.cell(row=curr, column=5, value=f"=SUM($H${first_line}:$H${last_item_num})")
            # 9. Col O
            ws.cell(row=curr, column=15, value=f"=SUM($H${first_line}:$H${last_item_num})")
            # 10. Col P
            ws.cell(row=curr, column=16, value=f"=$R${first_line}/($S${first_line}-900+$V${first_line})")
            # 11. Col R: Yellow + GBP
            cell_r = ws.cell(row=curr, column=18, value=gbp_paid)
            cell_r.fill = yellow_fill
            # 12. Col S: Yellow + YEN
            cell_s = ws.cell(row=curr, column=19, value=yen_paid)
            cell_s.fill = yellow_fill
            
            # 13. Col T: Pastel Orange + Formula
            cell_t = ws.cell(row=curr, column=20, value=f"=SUM($G${first_line}:$G${last_item_num})")
            cell_t.fill = pastel_orange_fill
            # 14. Col U: Pastel Orange + Formula
            cell_u = ws.cell(row=curr, column=21, value=f"=SUM($F${first_line}:$F${last_item_num})")
            cell_u.fill = pastel_orange_fill
            
            # 15. Col V
            cell_v = ws.cell(row=curr, column=22, value=f"=$T${first_line}-$U${first_line}")
            cell_v.fill = pastel_orange_fill

    # Footer: Add 'updated on' after the last item in Column B
    ws.cell(row=last_item_num + 1, column=2, value=f"updated on {order_date}")

    file_name = f"Suruga_Order_{first_line}.xlsx"
    wb.save(file_name)
    print(f"\nSuccess! Processed {len(items)} items.")
    print(f"File saved as: {file_name}")

if __name__ == "__main__":
    run_app()